import tkinter as tk
import tkinter.ttk as ttk
from openpyxl import *
from tkinter.messagebox import showinfo


def save():
    f_name = entry.get()
    l_name = entry1.get()
    age = entry2.get()

    wb = Workbook()
    ws = wb.active

    ws['A1'] = "First Name"
    ws['B1'] = "Last Name"
    ws['C1'] = "Age"
    ws['A2'] = f_name
    ws['B2'] = l_name
    ws['C2'] = age

    wb.save(r'C:\Teste\Python\dados.xlsx')
    showinfo("Saved", "Your entry has been saved!")


def clear():
    entry.delete(0, tk.END)
    entry1.delete(0, tk.END)
    entry2.delete(0, tk.END)


win = tk.Tk()
win.title("Registration Form")

label = tk.Label(win, text="First name: ")
label.grid(row=0, column=0, padx=8, pady=8)

entry = tk.Entry(win)
entry.grid(row=0, column=1, padx=8, pady=8)

label1 = tk.Label(win, text="Last name: ")
label1.grid(row=1, column=0, padx=8, pady=8)

entry1 = tk.Entry(win)
entry1.grid(row=1, column=1, padx=8, pady=8)

label2 = tk.Label(win, text="Age: ")
label2.grid(row=2, column=0, padx=8, pady=8)

entry2 = tk.Entry(win)
entry2.grid(row=2, column=1, padx=8, pady=8)

button = ttk.Button(win, text='Register', command=save)
button.grid(row=3, column=0, padx=8, pady=8)

button2 = ttk.Button(win, text='Clear', command=clear)
button2.grid(row=3, column=1, padx=8, pady=8)

win.mainloop()